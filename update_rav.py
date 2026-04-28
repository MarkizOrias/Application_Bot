#!/usr/bin/env python3
"""
update_rav.py — Sync cv/tracker.xlsx with RAV work-efforts on job-room.ch.

Flow:
  1. Read tracker, geocode missing Country / Post Code / City via Claude.
  2. Save enriched tracker.
  3. Open Chromium (visible) → EIAM login page.
  4. Wait for user to complete app-based authentication (up to 5 min).
  5. Navigate to work-efforts, scrape existing entries.
  6. For each tracker row:
       • Not yet on website  → submit new work-effort form.
       • Already on website  → update status if tracker differs.

Tracker column layout (1-based):
  1  Date          2  Country       3  Post Code     4  City
  5  Company       6  Job Title     7  URL           8  Folder
  9  Status        10 STATUS CHANGE
"""

import json
import os
import re
import time
from datetime import datetime
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import Page, sync_playwright, TimeoutError as PlaywrightTimeout

load_dotenv()

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

TRACKER_PATH = Path("cv/tracker.xlsx")

LOGIN_URL = "https://www.job-room.ch/dashboard/job-seeker"
WORK_EFFORTS_URL = "https://www.job-room.ch/work-efforts"

# Tracker column indices (1-based, openpyxl)
C_DATE = 1
C_COUNTRY = 2
C_POST_CODE = 3
C_CITY = 4
C_COMPANY = 5
C_JOB_TITLE = 6
C_URL = 7
C_FOLDER = 8
C_STATUS = 9
C_STATUS_CHANGE = 10

# Tracker status → German website label
TRACKER_TO_SITE = {
    "APPLIED": "Noch offen",
    "INTERVIEW": "Vorstellungsgespräch",
    "REJECTED": "Absage",
}

# Normalise Claude's country output to the exact English label used in the dropdown.
# The site's <select> uses English labels (e.g. "United States of America", "United Kingdom").
COUNTRY_LABEL_MAP: dict[str, str] = {
    "united states": "United States of America",
    "united states of america": "United States of America",
    "usa": "United States of America",
    "us": "United States of America",
    "uk": "United Kingdom",
    "great britain": "United Kingdom",
    "china": "People's Republic of China",
    "iran": "Islamic Republic of Iran",
    "north korea": "North Korea",
    "south korea": "South Korea",
    "taiwan": "Taiwan, Province of China",
    "moldova": "Moldova, Republic of",
    "north macedonia": "The Republic of North Macedonia",
    "tanzania": "United Republic of Tanzania",
    "remote": "Switzerland",  # fully remote → use Switzerland as fallback
}


# ---------------------------------------------------------------------------
# Tracker helpers
# ---------------------------------------------------------------------------

def read_tracker() -> tuple:
    """Return (workbook, worksheet, list_of_row_dicts)."""
    wb = load_workbook(TRACKER_PATH)
    ws = wb.active
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rows.append({
            "row_idx": i,
            "date": row[C_DATE - 1],
            "country": row[C_COUNTRY - 1],
            "post_code": row[C_POST_CODE - 1],
            "city": row[C_CITY - 1],
            "company": row[C_COMPANY - 1] or "",
            "job_title": row[C_JOB_TITLE - 1] or "",
            "url": row[C_URL - 1] or "",
            "folder": row[C_FOLDER - 1] or "",
            "status": (row[C_STATUS - 1] or "APPLIED").upper(),
            "status_change": row[C_STATUS_CHANGE - 1],
        })
    return wb, ws, rows


def save_tracker(wb, path: Path = TRACKER_PATH) -> None:
    wb.save(path)
    print(f"  [tracker] Saved → {path.resolve()}")


# ---------------------------------------------------------------------------
# Geocoding via Claude
# ---------------------------------------------------------------------------

def geocode_company(company: str, url: str, client: anthropic.Anthropic) -> dict:
    """Return company HQ country (English), post_code, city using Claude's knowledge."""
    prompt = (
        "You are a business-location lookup tool. Given a company name and job URL, "
        "return the company's HEADQUARTERS location as a JSON object.\n\n"
        f"Company: {company}\n"
        f"URL: {url or '(not available)'}\n\n"
        "Return ONLY a JSON object — no markdown, no explanation — with exactly these keys:\n"
        '  "country"   : country in English (e.g. "Switzerland", "United States")\n'
        '  "post_code" : HQ postal code as string (e.g. "1201", "10001"); '
        '"N/A" only if truly unknown\n'
        '  "city"      : HQ city (e.g. "Geneva", "New York")\n\n'
        "If the company is fully remote with no clear HQ, use its legal registration country/city."
    )

    resp = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=150,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.DOTALL).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        print(f"    [geocode] Unparseable response for '{company}': {raw!r}")
        return {"country": "Switzerland", "post_code": "N/A", "city": "N/A"}


def fill_missing_locations(wb, ws, rows: list, client: anthropic.Anthropic) -> None:
    """Geocode all rows that lack country/post_code/city and write back to workbook."""
    needs = [r for r in rows if not (r["country"] and r["post_code"] and r["city"])]
    if not needs:
        print("  [geocode] All entries already have location data — skipping.")
        return

    print(f"  [geocode] Enriching {len(needs)} entries…")
    for r in needs:
        loc = geocode_company(r["company"], r["url"], client)
        country = loc.get("country") or "Switzerland"
        post_code = str(loc.get("post_code") or "N/A")
        city = loc.get("city") or "N/A"

        ws.cell(r["row_idx"], C_COUNTRY).value = country
        ws.cell(r["row_idx"], C_POST_CODE).value = post_code
        ws.cell(r["row_idx"], C_CITY).value = city

        # Keep in-memory dicts current
        r["country"] = country
        r["post_code"] = post_code
        r["city"] = city

        print(f"    {r['company']:40s} → {country}, {post_code}, {city}")

    save_tracker(wb)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip().lower()


def _country_label(country_en: str) -> str:
    """Return the exact label text used in the site's country <select>."""
    return COUNTRY_LABEL_MAP.get(country_en.strip().lower(), country_en)


def parse_date_ddmmyyyy(value) -> str:
    """Convert any date-ish value to DD.MM.YYYY string."""
    if value is None:
        return datetime.now().strftime("%d.%m.%Y")
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return s


# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------

def wait_for_login(page: Page, timeout_ms: int = 300_000) -> None:
    # Give the page a moment to settle after navigation
    page.wait_for_load_state("networkidle")
    time.sleep(1)

    # Check whether we're already authenticated (profile/avatar visible)
    # vs needing to click the Login button
    already_in = (
        page.locator(
            "alv-user-info, "
            "[class*='user-info'], "
            "[class*='user-menu'], "
            "[class*='avatar'], "
            "button[aria-label*='profil'], "
            "button[aria-label*='Profil']"
        ).count() > 0
    )

    if already_in:
        print("  [login] Already authenticated — continuing.\n")
        return

    # Click the Login button
    login_btn = page.locator(
        "button:has-text('Login'), "
        "a:has-text('Login'), "
        "button:has-text('Anmelden'), "
        "a:has-text('Anmelden')"
    ).first
    login_btn.click()

    print(
        "\n  ┌─────────────────────────────────────────────────────────┐\n"
        "  │  Please log in using your Authenticator app.            │\n"
        "  │  Waiting up to 5 minutes…                               │\n"
        "  └─────────────────────────────────────────────────────────┘\n"
    )
    # Wait until redirected back to job-room.ch dashboard after EIAM auth
    page.wait_for_url("https://www.job-room.ch/**", timeout=timeout_ms)
    page.wait_for_load_state("networkidle")
    print("  [login] Authenticated — continuing.\n")


# ---------------------------------------------------------------------------
# Scrape existing work-efforts
# ---------------------------------------------------------------------------

def scrape_work_efforts(page: Page) -> list[dict]:
    """
    Return a list of dicts:
      { company, job_title, status, raw_text, element }

    job-room.ch is an Angular app; we try several selector strategies.
    """
    page.goto(WORK_EFFORTS_URL)
    page.wait_for_load_state("networkidle")
    time.sleep(2)

    # Angular component selectors (most specific → most generic)
    strategies = [
        "alv-work-effort-item",
        "[class*='work-effort-item']",
        "[class*='work-effort-card']",
        "mat-list-item",
        ".list-item",
        "tbody tr",
    ]
    cards = []
    for sel in strategies:
        found = page.locator(sel).all()
        if found:
            cards = found
            print(f"  [scrape] Using selector '{sel}' — {len(cards)} elements.")
            break

    entries: list[dict] = []
    for card in cards:
        try:
            raw = card.inner_text(timeout=2000)
        except Exception:
            raw = ""
        entry: dict = {
            "company": "",
            "job_title": "",
            "status": "",
            "raw_text": raw,
            "element": card,
        }
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        if lines:
            entry["company"] = lines[0]
        if len(lines) > 1:
            entry["job_title"] = lines[1]

        raw_low = raw.lower()
        if "vorstellungsgespräch" in raw_low:
            entry["status"] = "INTERVIEW"
        elif "absage" in raw_low:
            entry["status"] = "REJECTED"
        elif "noch offen" in raw_low:
            entry["status"] = "APPLIED"
        elif "anstellung" in raw_low:
            entry["status"] = "HIRED"

        entries.append(entry)

    print(f"  [scrape] Parsed {len(entries)} work-effort entries from the page.")
    return entries


def find_entry(site_entries: list[dict], company: str, job_title: str) -> dict | None:
    nc = _norm(company)
    nt = _norm(job_title)
    for e in site_entries:
        if _norm(e["company"]) == nc and _norm(e["job_title"]) == nt:
            return e
        # Partial: both normalised strings appear in the raw card text
        raw = _norm(e["raw_text"])
        if nc in raw and nt in raw:
            return e
    return None


# ---------------------------------------------------------------------------
# Create a new work-effort entry
# ---------------------------------------------------------------------------

# IDs are stable prefixes; the trailing "-0" suffix is dynamic but consistent.
_DATE_ID      = "input[id^='alv-date-input-portal.global.date']"
_ELEKTRONISCH = "input[id^='alv-checkbox-portal.work-efforts.edit-form.apply-channel.electronic']"
_COMPANY_ID   = "input[id^='alv-input-field-home.tools.job-publication.company.name']"
_COUNTRY_ID   = "select[id^='alv-select-home.tools.job-publication.company.country']"
_POSTCODE_ID  = "input[id^='alv-input-field-home.tools.job-publication.locality.zip']"
_CITY_ID      = "input[id^='alv-input-field-global.address.city']"
_JOBTITLE_ID  = "input[id^='alv-input-field-portal.global.job-title']"
_URL_ID       = "input[id^='alv-input-field-portal.work-efforts.edit-form.company.online-form-url']"
_RAV_NO       = "label[for$='-false']:has-text('No')"
_FULLTIME     = "label[for$='-FULLTIME']"
_PENDING      = "label[for$='-PENDING']"


def create_new_entry(page: Page, row: dict) -> None:
    page.goto(WORK_EFFORTS_URL)
    page.wait_for_load_state("networkidle")
    time.sleep(1)

    page.locator("a.add-work-effort-button").first.click()
    page.wait_for_load_state("networkidle")
    time.sleep(1)

    date_str = parse_date_ddmmyyyy(row["date"])

    # ── 1. Date ───────────────────────────────────────────────────────────
    date_inp = page.locator(_DATE_ID).first
    date_inp.click()
    date_inp.press("Control+a")
    date_inp.press_sequentially(date_str)
    page.keyboard.press("Escape")

    # ── 2. Elektronisch (checkbox) ────────────────────────────────────────
    page.locator(_ELEKTRONISCH).first.check()

    # ── 3. Company name ───────────────────────────────────────────────────
    page.locator(_COMPANY_ID).first.fill(row["company"])

    # ── 4. Country (native <select>, English labels) ──────────────────────
    page.locator(_COUNTRY_ID).first.select_option(
        label=_country_label(row["country"] or "Switzerland")
    )

    # ── 5. Postal code + City (separate inputs) ───────────────────────────
    post_code = str(row["post_code"] or "").replace("N/A", "").strip()
    if post_code:
        page.locator(_POSTCODE_ID).first.fill(post_code)

    city = str(row["city"] or "").strip()
    if city and city != "N/A":
        page.locator(_CITY_ID).first.fill(city)

    # ── 6. Job title ──────────────────────────────────────────────────────
    page.locator(_JOBTITLE_ID).first.fill(row["job_title"])

    # ── 7. URL ────────────────────────────────────────────────────────────
    url = str(row["url"] or "").strip()
    if url:
        page.locator(_URL_ID).first.fill(url)

    # ── 8. RAV Zuweisung → No ─────────────────────────────────────────────
    page.locator(_RAV_NO).first.click()

    # ── 9. Arbeitspensum → Vollzeit ───────────────────────────────────────
    page.locator(_FULLTIME).first.click()

    # ── 10. Ergebnis → Noch offen ─────────────────────────────────────────
    page.locator(_PENDING).first.click()

    # ── Save ──────────────────────────────────────────────────────────────
    page.locator(
        "button[type='submit'], button:has-text('Save'), button:has-text('Speichern')"
    ).first.click()
    page.wait_for_load_state("networkidle")
    time.sleep(1)
    print(f"    [new] Created: {row['company']} — {row['job_title']}")


# ---------------------------------------------------------------------------
# Update status of existing entry
# ---------------------------------------------------------------------------

def update_entry_status(
    page: Page, entry: dict, tracker_status: str, status_change_date
) -> None:
    site_label = TRACKER_TO_SITE.get(tracker_status)
    if not site_label:
        return

    # Click the entry card to open its detail view
    try:
        entry["element"].click()
        page.wait_for_load_state("networkidle")
        time.sleep(1)
    except Exception as e:
        print(f"    [warn] Could not click entry: {e}")
        return

    # If detail is read-only, look for an edit button
    edit_btn = page.locator(
        "button:has-text('Bearbeiten'), "
        "button[aria-label*='edit'], "
        "button[aria-label*='Bearbeiten']"
    ).first
    if edit_btn.is_visible(timeout=2000):
        edit_btn.click()
        page.wait_for_load_state("networkidle")
        time.sleep(1)

    # Change the Ergebnis radio button by clicking its label
    page.locator(f"label:has-text('{site_label}')").first.click()

    # Rejection popup: fill Absagegrund
    if site_label == "Absage":
        date_formatted = parse_date_ddmmyyyy(status_change_date)
        reason = (
            f"E-Mail vom {date_formatted}, "
            "die Personalabteilung hat einen anderen Kandidaten eingestellt."
        )
        try:
            page.wait_for_selector(
                "input[placeholder*='Absagegrund'], "
                "textarea[placeholder*='Absagegrund'], "
                "[formcontrolname*='rejectionReason'], "
                "[formcontrolname*='absagegrund']",
                timeout=4000,
            )
            page.locator(
                "input[placeholder*='Absagegrund'], "
                "textarea[placeholder*='Absagegrund'], "
                "[formcontrolname*='rejectionReason'], "
                "[formcontrolname*='absagegrund']"
            ).first.fill(reason)
        except PlaywrightTimeout:
            print("    [warn] Absagegrund field not found in popup.")

    # Save
    save_btn = page.locator(
        "button:has-text('Speichern'), "
        "button:has-text('Sichern'), "
        "button[type='submit']"
    ).first
    save_btn.click()
    page.wait_for_load_state("networkidle")
    time.sleep(1)
    print(f"    [status] Updated {entry['company']} → {site_label}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("RAV Work-Efforts Sync")
    print("=" * 60)

    if not TRACKER_PATH.exists():
        print(f"[!] Tracker not found: {TRACKER_PATH.resolve()}")
        return

    # Step 1: Read tracker
    wb, ws, rows = read_tracker()
    print(f"\n[1] Tracker loaded — {len(rows)} application entries.")

    # Step 2: Geocode missing locations
    api_key = os.environ.get("API_KEY") or os.environ.get("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)
    print("\n[2] Geocoding missing Country / Post Code / City via Claude…")
    fill_missing_locations(wb, ws, rows, client)

    # Step 3: Open browser
    print("\n[3] Opening browser for EIAM / RAV login…")
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=80)
        ctx = browser.new_context()
        page = ctx.new_page()

        page.goto(LOGIN_URL)
        wait_for_login(page)

        # Step 4: Scrape existing entries
        print("[4] Scraping existing work-efforts from job-room.ch…")
        site_entries = scrape_work_efforts(page)

        # Step 5: Sync
        print(f"\n[5] Syncing {len(rows)} tracker entries…\n")
        created = updated = skipped = 0

        for row in rows:
            company = row["company"]
            title = row["job_title"]
            tracker_status = row["status"]

            try:
                existing = find_entry(site_entries, company, title)

                if existing is None:
                    print(f"  → NEW    : {company} — {title}")
                    create_new_entry(page, row)
                    created += 1
                else:
                    site_status = existing["status"]
                    if site_status != tracker_status:
                        print(
                            f"  → UPDATE : {company} — {title}"
                            f"  (tracker={tracker_status}, site={site_status})"
                        )
                        update_entry_status(
                            page, existing, tracker_status, row["status_change"]
                        )
                        updated += 1
                    else:
                        print(f"  → OK     : {company} — {title}  [{tracker_status}]")
                        skipped += 1
            except Exception as e:
                import traceback
                print(f"    [error] {company} — {title}: {e}")
                traceback.print_exc()

        print(
            f"\n  Summary: {created} created, {updated} updated, {skipped} already in sync."
        )
        input("\n  Press Enter to close the browser…")
        browser.close()

    print("\nAll done.")


if __name__ == "__main__":
    main()
