#!/usr/bin/env python3
"""
LinkedIn Job Scraper - Phase 1 of Application Bot
Automatically kills any running Chrome, launches a fresh instance with remote
debugging on port 9222, attaches via CDP, and saves job listings to Excel.

Usage:
    python linkedin_scraper.py

The script handles Chrome startup automatically. Log in to LinkedIn if prompted.
"""

import json
import re
import subprocess
import time
from datetime import datetime
from pathlib import Path

from cv_generator import generate_cvs_for_jobs
from apply_bot import (
    load_tracker,
    upsert_jobs,
    mark_applied,
    already_applied,
    save_tracker,
    attempt_easy_apply,
    load_mismatch_log,
    save_mismatch,
)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

CONFIG_PATH = Path("config/profile.json")
OUTPUT_DIR = Path("output")

# LinkedIn brand blue for header styling
LI_BLUE = "0A66C2"


CHROME_EXE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROME_USER_DATA = r"C:\Temp\chrome-debug"
CDP_PORT = 9222
CDP_URL = f"http://localhost:{CDP_PORT}"


# ---------------------------------------------------------------------------
# Chrome lifecycle
# ---------------------------------------------------------------------------


def _wait_for_cdp(timeout: float = 15.0) -> bool:
    """Poll the CDP endpoint until Chrome exposes it. Returns True when ready."""
    import urllib.request
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            urllib.request.urlopen(f"{CDP_URL}/json", timeout=1)
            return True
        except Exception:
            time.sleep(0.4)
    return False


def launch_chrome() -> None:
    """Kill all Chrome processes, then start a fresh instance with CDP enabled."""
    print("Killing existing Chrome processes...")
    subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"], capture_output=True)
    time.sleep(0.5)  # brief wait for process cleanup

    print("Launching Chrome with remote debugging...")
    subprocess.Popen(
        [
            CHROME_EXE,
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={CHROME_USER_DATA}",
            "--start-maximized",
        ]
    )

    if _wait_for_cdp():
        print("    Chrome ready.")
    else:
        print("[!] Chrome did not expose CDP within 15s — proceeding anyway")


# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------


def load_profile() -> dict:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# URL building
# ---------------------------------------------------------------------------


def build_search_url(role: str, easy_apply: bool) -> str:
    """Build a LinkedIn jobs search URL for a given role."""
    from urllib.parse import urlencode

    params = {
        "keywords": role,
        "f_WT": "2",  # Remote only
        "sortBy": "DD",  # Most recent first
    }
    if easy_apply:
        params["f_AL"] = "true"
    return "https://www.linkedin.com/jobs/search/?" + urlencode(params)


# ---------------------------------------------------------------------------
# Scraping helpers
# ---------------------------------------------------------------------------


def scroll_jobs_panel(page, max_rounds: int = 8) -> None:
    """Scroll the jobs list panel until card count stabilises or max_rounds reached."""
    panel = (
        page.query_selector(".jobs-search-results-list")
        or page.query_selector(".scaffold-layout__list-container")
        or page.query_selector("[class*='jobs-search-results-list']")
    )
    prev_count = 0
    for _ in range(max_rounds):
        if panel:
            panel.evaluate("el => el.scrollBy(0, 800)")
        else:
            page.evaluate("window.scrollBy(0, 800)")
        time.sleep(0.6)
        cards = page.query_selector_all("li[data-occludable-job-id]")
        if len(cards) == prev_count and prev_count > 0:
            break  # no new cards loaded — stop early
        prev_count = len(cards)


def extract_card(card) -> dict | None:
    """Extract fields from a single job-card element. Returns None on failure."""

    # Job ID is the most reliable field — present as a data attribute on the <li>
    job_id = (
        card.get_attribute("data-occludable-job-id")
        or card.get_attribute("data-job-id")
        or card.get_attribute(
            "data-entity-urn",
        )
    )

    # Title link — try stable selectors first, then fall back to any /jobs/view/ href
    title_el = (
        card.query_selector("a.job-card-list__title--link")
        or card.query_selector("a[href*='/jobs/view/']")
        or card.query_selector("a.job-card-container__link")
    )

    company_el = (
        card.query_selector(".job-card-container__company-name")
        or card.query_selector(".artdeco-entity-lockup__subtitle span")
        or card.query_selector(".job-card-container__primary-description")
    )

    location_el = (
        card.query_selector("li.job-card-container__metadata-item")
        or card.query_selector(".job-card-container__metadata-item")
        or card.query_selector(".artdeco-entity-lockup__caption li")
    )

    title = title_el.inner_text().strip() if title_el else None
    if not title:
        return None

    company = company_el.inner_text().strip() if company_el else ""
    location = location_el.inner_text().strip() if location_el else ""

    # Build clean job URL
    if job_id:
        # Strip urn prefix if present  e.g. "urn:li:job:123" -> "123"
        job_id = re.sub(r".*:", "", str(job_id))
        clean_url = f"https://www.linkedin.com/jobs/view/{job_id}/"
    else:
        href = title_el.get_attribute("href") if title_el else ""
        if href and not href.startswith("http"):
            href = "https://www.linkedin.com" + href
        m = re.search(r"/jobs/view/(\d+)", href or "")
        clean_url = f"https://www.linkedin.com/jobs/view/{m.group(1)}/" if m else href

    return {
        "title": title,
        "company": company,
        "location": location,
        "url": clean_url,
        "easy_apply": False,  # filled in by _enrich_from_panel
        "description": None,  # filled in by _enrich_from_panel
        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "applied": False,
    }


def _enrich_from_panel(page, card, job: dict) -> None:
    """Click a job card to load the right-hand detail panel, then read the
    description and apply-button type without navigating away from the search page."""
    try:
        link = card.query_selector("a[href*='/jobs/view/']")
        if link:
            link.click()
        else:
            card.click()

        # Wait for the detail panel to populate
        try:
            page.wait_for_selector("#job-details", timeout=8000)
        except Exception:
            return

        time.sleep(0.4)

        # Read job description from the panel
        desc_el = page.query_selector("#job-details")
        if desc_el:
            text = desc_el.inner_text().strip()
            if len(text) > 100:
                job["description"] = text[:4500]

        # Check for "Applied X ago" feedback message — skip if already applied
        for el in page.query_selector_all(".artdeco-inline-feedback__message"):
            if "applied" in (el.inner_text() or "").lower():
                job["linkedin_applied"] = True
                return  # no need to read the rest

        # Detect apply type from the panel button
        # Easy Apply: <button id="jobs-apply-button-id" aria-label*="Easy Apply">
        easy_btn = page.query_selector(
            "button#jobs-apply-button-id"
        ) or page.query_selector("button[aria-label*='Easy Apply']")
        job["easy_apply"] = bool(easy_btn)

    except Exception as exc:
        print(f"    [warn] Panel read error for '{job.get('title')}': {exc}")


def scrape_cards(page, max_per_search: int = 25) -> list[dict]:
    """Wait for cards, scroll to load more, click each to enrich with description
    and apply-type from the right-hand panel, then return all jobs."""
    jobs = []

    card_selector = (
        "li[data-occludable-job-id], "
        "li[data-job-id], "
        "li.jobs-search-results__list-item"
    )

    try:
        page.wait_for_selector(card_selector, timeout=18000)
    except PlaywrightTimeout:
        print("    [warn] No job cards found - check login or CAPTCHA")
        return jobs

    scroll_jobs_panel(page)

    cards = (
        page.query_selector_all("li[data-occludable-job-id]")
        or page.query_selector_all("li[data-job-id]")
        or page.query_selector_all("li.jobs-search-results__list-item")
    )

    print(f"    Found {len(cards)} cards on page")

    for card in cards[:max_per_search]:
        try:
            job = extract_card(card)
            if not job:
                continue
            _enrich_from_panel(page, card, job)
            if job.get("linkedin_applied"):
                print(
                    f"    [skip] Already applied on LinkedIn: {job['title']} @ {job['company']}"
                )
                continue
            jobs.append(job)
        except Exception as exc:
            print(f"    [warn] Card extraction error: {exc}")

    return jobs


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------


def is_excluded(job: dict, profile: dict) -> tuple[bool, str]:
    """Return (True, reason) if the job should be skipped, else (False, '')."""
    prefs = profile["preferences"]

    # Excluded companies (case-insensitive)
    blocked = {c.lower() for c in prefs.get("exclude_companies", [])}
    if job["company"].lower() in blocked:
        return True, f"excluded company '{job['company']}'"

    # Keyword exclusions checked against title
    title_lower = job["title"].lower()
    for kw in prefs.get("exclude_keywords", []):
        if kw.lower() in title_lower:
            return True, f"exclude keyword '{kw}'"

    return False, ""


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def save_excel(jobs: list[dict], path: Path) -> None:
    """Write jobs list to a formatted Excel file."""
    df = pd.DataFrame(jobs)

    # Column order
    col_order = [
        "title",
        "company",
        "location",
        "search_role",
        "easy_apply",
        "url",
        "applied",
        "scraped_at",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    df.to_excel(path, index=False, sheet_name="Listings")

    # --- Formatting with openpyxl ---
    wb = load_workbook(path)
    ws = wb.active

    # Header row styling
    header_fill = PatternFill(start_color=LI_BLUE, end_color=LI_BLUE, fill_type="solid")
    header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Make URL column a hyperlink and style body rows
    url_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).lower() == "url":
            url_col_idx = idx
            break

    body_font = Font(name="Calibri", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if url_col_idx and cell.column == url_col_idx and cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(
                    name="Calibri", size=10, color="0563C1", underline="single"
                )
                cell.value = "Open"

    # Auto-size columns
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=10,
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"\nSaved {len(jobs)} listings -> {path.resolve()}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    profile = load_profile()
    prefs = profile["preferences"]
    li_cfg = profile["job_sources"]["linkedin"]

    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = OUTPUT_DIR / f"linkedin_listings_{timestamp}.xlsx"

    easy_apply = li_cfg.get("easy_apply_only", True)

    print("=" * 60)
    print("LinkedIn Job Scraper")
    print("=" * 60)
    print(f"Easy Apply only: {easy_apply}")
    print(f"Roles to search: {len(prefs['roles'])}")
    print(f"Output         : {output_path}")
    print("=" * 60)
    print()

    launch_chrome()

    all_jobs: list[dict] = []
    seen_urls: set[str] = set()

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.connect_over_cdp(CDP_URL)
        except Exception as exc:
            print(f"[!] Could not connect to Chrome: {exc}")
            return

        # Reuse the existing context (your logged-in session)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.new_page()

        # Warm-up: make sure we're logged in
        print("Opening LinkedIn...")
        page.goto(
            "https://www.linkedin.com/feed/",
            wait_until="domcontentloaded",
            timeout=30000,
        )
        time.sleep(2)

        if "login" in page.url or "authwall" in page.url:
            print("\n[!] LinkedIn is asking for login.")
            print(
                "    Please log in manually in the browser, then press Enter here to continue."
            )
            input("    Press Enter once logged in > ")

        df = load_tracker()
        mismatched_urls = load_mismatch_log()
        max_apps = profile["preferences"].get("max_applications_per_session", 15)
        applied_count = 0

        # --- Search → CV → apply pipeline (one card at a time) ---
        for role in prefs["roles"]:
            if applied_count >= max_apps:
                print(f"\n[apply] Session limit reached ({max_apps}).")
                break

            search_url = build_search_url(role, easy_apply)
            print(f"\nSearching: {role}")
            print(f"  URL: {search_url}")

            try:
                page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                time.sleep(2)

                cards = scrape_cards(page)
                new_count = 0

                for job in cards:
                    if applied_count >= max_apps:
                        print(f"\n[apply] Session limit reached ({max_apps}).")
                        break

                    job_url = job["url"]

                    # De-duplicate by URL
                    if job_url in seen_urls:
                        continue
                    seen_urls.add(job_url)

                    excluded, reason = is_excluded(job, profile)
                    if excluded:
                        print(
                            f"  [-] Skip  : {job['title']} @ {job['company']} ({reason})"
                        )
                        continue

                    job["search_role"] = role
                    all_jobs.append(job)
                    new_count += 1
                    print(
                        f"  [+] Found : {job['title']} @ {job['company']} | {job['location']}"
                    )

                    # --- Immediately process this job ---
                    if already_applied(df, job_url):
                        print(
                            f"  [skip] Already applied: {job['title']} @ {job['company']}"
                        )
                        continue

                    if job_url in mismatched_urls:
                        print(
                            f"  [skip] Previous mismatch: {job['title']} @ {job['company']}"
                        )
                        continue

                    cv_map, new_mismatches = generate_cvs_for_jobs(
                        [job], profile, page, limit=1
                    )
                    for m in new_mismatches:
                        save_mismatch(m["url"], m["title"], m["company"], m["reason"])
                        mismatched_urls.add(m["url"])
                    if not cv_map:
                        continue  # mismatch or render error — don't track

                    cv_path = cv_map.get(job_url)
                    df = upsert_jobs(df, [job])

                    if not job.get("easy_apply"):
                        print(
                            f"  [apply] External apply — see URL in tracker: {job['title']} @ {job['company']}"
                        )
                        continue

                    success = attempt_easy_apply(page, job, profile, cv_path)
                    if success:
                        df = mark_applied(df, job_url, cv_path)
                        applied_count += 1
                        print(f"    [apply] ✓ Applied ({applied_count}/{max_apps})")
                    time.sleep(2)

                print(f"  --> {new_count} new jobs found for '{role}'")

            except PlaywrightTimeout:
                print(f"  [!] Timeout navigating to search for '{role}' - skipping")
            except Exception as exc:
                print(f"  [!] Error on '{role}': {exc}")

            time.sleep(1.5)  # Polite delay between searches

        save_tracker(df)
        print(
            f"\n[apply] Session complete — {applied_count} application(s) submitted."
        )

        context.close()

    # --- Save raw listings snapshot (timestamped, read-only reference) ---
    if all_jobs:
        save_excel(all_jobs, output_path)
    else:
        print("\nNo jobs collected. Check selectors or LinkedIn login state.")


if __name__ == "__main__":
    main()
