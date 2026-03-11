#!/usr/bin/env python3
"""
apply_bot.py — LinkedIn Easy Apply automation + persistent job tracker.

Handles:
  • attempt_easy_apply()   — navigate the full Easy Apply modal for one job
  • run_apply_session()    — orchestrate apply loop with skip/limit logic
  • Tracker CRUD           — load / upsert / mark-applied / save Excel tracker
"""

import json
import os
import re
import time
from datetime import datetime
from pathlib import Path

import anthropic
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeout

load_dotenv()

TRACKER_PATH = Path("output/linkedin_tracker.xlsx")
MISMATCH_PATH = Path("output/mismatch_log.json")
LI_BLUE = "0A66C2"

TRACKER_COLS = [
    "title", "company", "location", "url",
    "easy_apply", "search_role", "scraped_at",
    "applied", "applied_at", "cv_path",
]

# ---------------------------------------------------------------------------
# Anthropic client — created once per process, reused across all form-fill calls
# ---------------------------------------------------------------------------

_anthropic_client: anthropic.Anthropic | None = None


def _get_client() -> anthropic.Anthropic:
    global _anthropic_client
    if _anthropic_client is None:
        api_key = os.environ.get("API_KEY") or os.environ.get("ANTHROPIC_API_KEY")
        _anthropic_client = anthropic.Anthropic(api_key=api_key)
    return _anthropic_client


# ===========================================================================
# Tracker management
# ===========================================================================

def load_tracker() -> pd.DataFrame:
    """Load existing tracker or return an empty DataFrame with correct columns."""
    if TRACKER_PATH.exists():
        try:
            df = pd.read_excel(TRACKER_PATH, sheet_name="Tracker")
            return df.reindex(columns=TRACKER_COLS)
        except Exception as exc:
            print(f"  [tracker] Could not read tracker ({exc}) — starting fresh.")
    return pd.DataFrame(columns=TRACKER_COLS)


def upsert_jobs(df: pd.DataFrame, jobs: list[dict]) -> pd.DataFrame:
    """Add jobs not already in the tracker. Existing rows are never overwritten."""
    existing = set(df["url"].dropna().tolist())
    new_rows = []
    for job in jobs:
        if job.get("url") in existing:
            continue
        row = {col: job.get(col) for col in TRACKER_COLS}
        row["applied"] = False
        row["applied_at"] = None
        row["cv_path"] = None
        new_rows.append(row)
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows, columns=TRACKER_COLS)], ignore_index=True)
        print(f"  [tracker] Added {len(new_rows)} new job(s).")
    return df


def mark_applied(df: pd.DataFrame, url: str, cv_path: Path | None = None) -> pd.DataFrame:
    """Mark a job as applied and record the timestamp and CV used."""
    mask = df["url"] == url
    if mask.any():
        df.loc[mask, "applied"] = True
        df.loc[mask, "applied_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        if cv_path:
            df.loc[mask, "cv_path"] = str(cv_path)
    return df


def already_applied(df: pd.DataFrame, url: str) -> bool:
    """Return True if the URL has been marked applied in the tracker."""
    rows = df.loc[df["url"] == url, "applied"]
    return bool(rows.any() and rows.iloc[0])


def save_tracker(df: pd.DataFrame) -> None:
    """Write tracker to Excel with LinkedIn-branded header formatting."""
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(TRACKER_PATH, index=False, sheet_name="Tracker")

    wb = load_workbook(TRACKER_PATH)
    ws = wb.active

    header_fill = PatternFill(start_color=LI_BLUE, end_color=LI_BLUE, fill_type="solid")
    header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Hyperlink URL column
    url_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).lower() == "url":
            url_col_idx = idx
            break

    body_font = Font(name="Calibri", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if url_col_idx and cell.column == url_col_idx and cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                cell.value = "Open"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"
    wb.save(TRACKER_PATH)
    print(f"  [tracker] Saved -> {TRACKER_PATH.resolve()}")


# ===========================================================================
# Mismatch log — jobs Claude determined are not a profile fit
# ===========================================================================

def load_mismatch_log() -> set[str]:
    """Return the set of job URLs previously marked as profile mismatches."""
    if MISMATCH_PATH.exists():
        try:
            data = json.loads(MISMATCH_PATH.read_text(encoding="utf-8"))
            return set(data.keys())
        except Exception:
            pass
    return set()


def save_mismatch(url: str, title: str, company: str, reason: str) -> None:
    """Append a mismatch entry to the persistent log (no-op if already present)."""
    MISMATCH_PATH.parent.mkdir(parents=True, exist_ok=True)
    data: dict = {}
    if MISMATCH_PATH.exists():
        try:
            data = json.loads(MISMATCH_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    if url in data:
        return
    data[url] = {
        "title": title,
        "company": company,
        "reason": reason,
        "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    MISMATCH_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"  [mismatch] Logged: {title} @ {company}")


# ===========================================================================
# Question-answering helpers
# ===========================================================================

def _exp_years(profile: dict) -> str:
    """Estimate total years of professional experience from the earliest job date."""
    all_years = []
    for e in profile.get("cv", {}).get("experience", []):
        for y in re.findall(r"\b(20\d{2}|19\d{2})\b", e.get("period", "")):
            all_years.append(int(y))
    if not all_years:
        return "7"
    return str(max(1, datetime.now().year - min(all_years)))



def _candidate_context(profile: dict) -> dict:
    """Build a compact candidate snapshot to send with every Claude request."""
    cv = profile.get("cv", {})
    p = profile["personal"]
    return {
        "full_name": p.get("full_name", ""),
        "email": p.get("email", ""),
        "phone": p.get("phone", ""),
        "location": p.get("location", ""),
        "linkedin": p.get("linkedin", ""),
        "github": p.get("github", ""),
        "total_experience_years": _exp_years(profile),
        "work_authorization": profile.get("work_authorization", {}),
        "min_salary_usd": profile["preferences"].get("min_salary_usd", 70000),
        "notice_period": "2 weeks",
        "skills": cv.get("skills", []),
        "experience": [
            {
                "title": e.get("title", ""),
                "company": e.get("company", ""),
                "period": e.get("period", ""),
                "bullets": e.get("bullets", []),
            }
            for e in (cv.get("experience") or [])[:5]
        ],
    }


# ===========================================================================
# Form scraping — collect all questions from the current modal step
# ===========================================================================

def _scrape_form_fields(page: Page) -> list[dict]:
    """Return a list of field descriptors for every visible, unfilled form field."""
    fields = []

    def _label_for(el) -> str:
        el_id = el.get_attribute("id") or ""
        if el_id:
            lbl = page.query_selector(f"label[for='{el_id}']")
            if lbl:
                return lbl.inner_text().strip()
        return el.evaluate("""el => {
            let node = el.parentElement;
            for (let i = 0; i < 7; i++) {
                if (!node) break;
                const lbl = node.querySelector('label, legend, [class*="label"]');
                if (lbl && lbl.innerText.trim().length > 1) return lbl.innerText.trim();
                node = node.parentElement;
            }
            return '';
        }""").strip()

    # --- text / number / tel / email inputs ---
    for inp in page.query_selector_all(
        "input[type='text'], input[type='number'], input[type='tel'], input[type='email']"
    ):
        try:
            if not inp.is_visible() or (inp.input_value() or "").strip():
                continue
            label = _label_for(inp)
            if not label:
                continue
            inp_id = inp.get_attribute("id") or label
            # LinkedIn uses type="text" with id ending in "-numeric" for whole-number fields
            is_numeric = (
                inp.get_attribute("type") == "number"
                or (inp_id or "").endswith("-numeric")
            )
            fields.append({
                "id": inp_id,
                "kind": "number" if is_numeric else "text",
                "label": label,
                "options": [],
            })
        except Exception:
            pass

    # --- textareas ---
    for ta in page.query_selector_all("textarea"):
        try:
            if not ta.is_visible() or (ta.input_value() or "").strip():
                continue
            label = _label_for(ta)
            if not label:
                continue
            ta_id = ta.get_attribute("id") or label
            fields.append({"id": ta_id, "kind": "textarea", "label": label, "options": []})
        except Exception:
            pass

    # --- selects ---
    for sel in page.query_selector_all("select"):
        try:
            if not sel.is_visible():
                continue
            current = (sel.input_value() or "").lower()
            if current and current not in ("", "select an option", "-- none --"):
                continue
            label = _label_for(sel)
            if not label:
                continue
            sel_id = sel.get_attribute("id") or label
            options = []
            for opt in page.query_selector_all(f"#{sel_id} option" if sel.get_attribute("id") else "option"):
                val = opt.get_attribute("value") or ""
                txt = opt.inner_text().strip()
                if val and txt and txt.lower() not in ("select an option", "-- none --", ""):
                    options.append(txt)
            fields.append({"id": sel_id, "kind": "select", "label": label, "options": options})
        except Exception:
            pass

    # --- radio groups ---
    radio_names: set[str] = set()
    for r in page.query_selector_all("input[type='radio']"):
        n = r.get_attribute("name")
        if n:
            radio_names.add(n)

    for name in radio_names:
        try:
            radios = page.query_selector_all(f"input[type='radio'][name='{name}']")
            if not radios or any(r.is_checked() for r in radios):
                continue
            safe_name = name.replace('"', '\\"')
            question = page.evaluate(f"""() => {{
                const r = document.querySelector('input[type="radio"][name="{safe_name}"]');
                if (!r) return '';
                let node = r.closest('fieldset') || r.parentElement;
                for (let i = 0; i < 7; i++) {{
                    if (!node) break;
                    const leg = node.querySelector('legend, [class*="label"]');
                    if (leg && leg.innerText.trim().length > 2) return leg.innerText.trim();
                    node = node.parentElement;
                }}
                return '';
            }}""")
            options = []
            for r in radios:
                rid = r.get_attribute("id")
                lbl_el = page.query_selector(f"label[for='{rid}']") if rid else None
                if lbl_el:
                    opt_text = lbl_el.inner_text().strip()
                else:
                    # LinkedIn renders toggle-style radios without <label> elements;
                    # option text is in the value or data attribute instead
                    opt_text = (
                        r.get_attribute("data-test-text-selectable-option__input")
                        or r.get_attribute("value")
                        or ""
                    ).strip()
                if opt_text:
                    options.append(opt_text)
            if options:
                fields.append({
                    "id": f"radio::{name}",
                    "kind": "radio",
                    "label": question or options[0],
                    "options": options,
                })
        except Exception:
            pass

    return fields


# ===========================================================================
# Claude-driven batch answer
# ===========================================================================

def _claude_fill_form(page: Page, profile: dict) -> None:
    """
    Scrape all unfilled fields from the current modal step, ask Claude Haiku
    to answer them all at once with full candidate context, then apply answers.
    """
    fields = _scrape_form_fields(page)
    if not fields:
        return

    # Build question list for the prompt
    q_lines = []
    for f in fields:
        opts = f" [options: {', '.join(f['options'])}]" if f["options"] else ""
        num_hint = " (integer 0-99)" if f["kind"] == "number" else ""
        q_lines.append(f'  "{f["id"]}": "{f["label"]}"{opts}{num_hint}')

    candidate = _candidate_context(profile)
    prompt = (
        "You are filling a LinkedIn Easy Apply form on behalf of this candidate.\n\n"
        f"CANDIDATE:\n{json.dumps(candidate, indent=2)}\n\n"
        "FORM FIELDS (id → question):\n"
        + "\n".join(q_lines)
        + "\n\n"
        "Return a JSON object where each key is a field id and the value is the answer.\n"
        "Rules:\n"
        "- For 'number' fields return only a digit string, e.g. \"5\"\n"
        "- For 'select' or 'radio' fields copy the answer verbatim from the listed options\n"
        "- For 'textarea' fields write 2-4 concise sentences appropriate for a job application\n"
        "- If a field is not relevant or unknown, return an empty string\n"
        "- Return ONLY valid JSON, no markdown fences, no extra text"
    )

    try:
        resp = _get_client().messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        # Strip accidental markdown fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        answers: dict = json.loads(raw)
    except Exception as exc:
        print(f"      [apply] Claude form fill error: {exc}")
        return

    print(f"      [apply] Claude answered {len(answers)} field(s)")

    # Apply answers back to the DOM
    for f in fields:
        answer = (answers.get(f["id"]) or "").strip()
        if not answer:
            continue
        try:
            kind = f["kind"]
            fid = f["id"]
            label = f["label"]

            if kind in ("text", "textarea", "number"):
                if kind == "number":
                    digits = re.sub(r"[^\d]", "", answer.split(".")[0])
                    answer = digits if digits else "0"
                # Use attribute selector [id="..."] so IDs with special CSS characters
                # (colons, URN format) work. locator.fill() simulates real keyboard input
                # which properly updates React controlled component state — unlike the
                # native value setter trick which leaves React's fiber state out of sync.
                loc = page.locator(f'[id="{fid}"]')
                if loc.count() > 0:
                    loc.first.click()
                    loc.first.fill(answer)
                    loc.first.dispatch_event("blur")

            elif kind == "select":
                sel_el = page.query_selector(f"#{fid}") or page.query_selector(f"[id='{fid}']")
                if sel_el and sel_el.is_visible():
                    matched = False
                    for txt in f["options"]:
                        if txt.lower() == answer.lower():
                            sel_el.select_option(label=txt)
                            matched = True
                            break
                    if not matched:
                        for txt in f["options"]:
                            if answer.lower() in txt.lower() or txt.lower() in answer.lower():
                                sel_el.select_option(label=txt)
                                break

            elif kind == "radio":
                radio_name = fid.replace("radio::", "", 1)
                # Use JS to match by label text, data attribute, or value — avoids
                # CSS selector issues with URN-style name attributes.
                page.evaluate(
                    """([radioName, answerLower]) => {
                        for (const r of document.querySelectorAll('input[type="radio"]')) {
                            if (r.name !== radioName) continue;
                            const lbl = r.id
                                ? document.querySelector('label[for="' + r.id + '"]')
                                : null;
                            const txt = (lbl
                                ? lbl.innerText
                                : (r.getAttribute('data-test-text-selectable-option__input')
                                   || r.value || '')
                            ).toLowerCase().trim();
                            if (txt === answerLower || txt.includes(answerLower)
                                    || answerLower.includes(txt)) {
                                r.click();
                                return true;
                            }
                        }
                        return false;
                    }""",
                    [radio_name, answer.lower()],
                )

        except Exception as exc:
            print(f"      [apply] apply answer error for '{label}': {exc}")


def _uncheck_optionals(page: Page) -> None:
    """Uncheck opt-in checkboxes for Follow, Featured/Top applicant, etc."""
    SPAM_PATTERNS = [
        "follow", "featured applicant", "top applicant", "top choice",
        "notify me", "receive email", "get email",
    ]
    for cb in page.query_selector_all("input[type='checkbox']"):
        try:
            if not cb.is_visible() or not cb.is_checked():
                continue
            cb_id = cb.get_attribute("id")
            lbl_el = page.query_selector(f"label[for='{cb_id}']") if cb_id else None
            label = (lbl_el.inner_text().strip() if lbl_el else cb.evaluate(
                "el => { let n=el.parentElement; for(let i=0;i<5;i++){if(!n)break;"
                "const l=n.querySelector('label,[class*=\"label\"]');"
                "if(l&&l.innerText.trim().length>1)return l.innerText.trim();n=n.parentElement;}return '';}"
            )).lower()
            if any(pat in label for pat in SPAM_PATTERNS):
                cb.click()
                print(f"      [apply] Unchecked: '{label[:60]}'")
        except Exception:
            pass


def _upload_resume(page: Page, cv_path: Path) -> bool:
    """Upload the tailored CV PDF. Works even on hidden file inputs."""
    try:
        file_input = page.query_selector("input[type='file']")
        if not file_input:
            return False
        file_input.set_input_files(str(cv_path.resolve()))
        time.sleep(1.5)
        print(f"      [apply] Uploaded: {cv_path.name}")
        return True
    except Exception as exc:
        print(f"      [apply] Upload failed: {exc}")
        return False


def _advance(page: Page) -> str:
    """Click Submit / Review / Next. Returns which button was clicked, or 'stuck'."""
    for aria_label, key in [
        ("Submit application", "submit"),
        ("Review your application", "review"),
        ("Continue to next step", "next"),
    ]:
        btn = page.query_selector(f"button[aria-label='{aria_label}']")
        if btn and btn.is_visible() and btn.is_enabled():
            btn.click()
            return key

    # Fallback: single JS pass over all buttons — avoids repeated Python↔DOM round-trips
    result = page.evaluate("""() => {
        const targets = [
            ["Submit application", "submit"],
            ["Review", "review"],
            ["Next", "next"],
            ["Continue", "next"],
        ];
        for (const [text, key] of targets) {
            for (const btn of document.querySelectorAll("button")) {
                if (!btn.offsetParent || btn.disabled) continue;
                const span = btn.querySelector("span");
                const t = (span ? span.innerText : btn.innerText).trim();
                if (t === text) { btn.click(); return key; }
            }
        }
        return "stuck";
    }""")
    return result or "stuck"


def _close_modal(page: Page) -> None:
    """Dismiss the Easy Apply modal and confirm discard if prompted."""
    for sel in ["button[aria-label='Dismiss']", "button[aria-label='Close']",
                ".artdeco-modal__dismiss"]:
        btn = page.query_selector(sel)
        if btn and btn.is_visible():
            try:
                btn.click()
                time.sleep(0.8)
            except Exception:
                pass
            break

    # Confirm discard dialog
    for discard_sel in [
        "button[data-test-dialog-primary-btn]",
        "button[aria-label='Discard']",
        "button[data-control-name='discard_application_confirm_btn']",
    ]:
        btn = page.query_selector(discard_sel)
        if btn and btn.is_visible():
            try:
                btn.click()
            except Exception:
                pass
            return

    # Last resort: any visible "Discard" button
    for btn in page.query_selector_all("button"):
        try:
            if btn.is_visible() and "discard" in (btn.inner_text() or "").lower():
                btn.click()
                return
        except Exception:
            pass


# ===========================================================================
# Core modal handler
# ===========================================================================

def _run_modal(page: Page, profile: dict, cv_path: Path | None) -> bool:
    """Step through the Easy Apply modal. Returns True if successfully submitted."""
    MODAL_SEL = "div.jobs-easy-apply-modal, [data-test-modal-id='easy-apply-modal']"
    MAX_STEPS = 12
    cv_uploaded = False
    last_result: str = ""
    repeat_count: int = 0

    for step in range(MAX_STEPS):
        time.sleep(0.6)

        if not page.query_selector(MODAL_SEL):
            print("      [apply] Modal closed unexpectedly")
            return False

        # On the submit step there are no fields to fill — skip scanning.
        # NOTE: "Review your application" button appears on the LAST form step
        # (e.g. Additional Questions) as the advance button, so we must NOT treat
        # that step as a review step — only the final Submit page is skipped.
        is_review = bool(
            page.query_selector("button[aria-label='Submit application']")
        )

        if not is_review:
            # Upload CV once (first step that has a file input)
            if cv_path and not cv_uploaded:
                if page.query_selector("input[type='file']"):
                    # Temporarily rename to FirstName_LastName.pdf so the recruiter
                    # sees a clean name; rename back to company_position after upload
                    name_slug = "_".join(
                        profile["personal"].get("full_name", "Candidate").split()
                    )
                    temp_path = cv_path.parent / f"{name_slug}.pdf"
                    cv_path.rename(temp_path)
                    try:
                        cv_uploaded = _upload_resume(page, temp_path)
                    finally:
                        temp_path.rename(cv_path)

            _uncheck_optionals(page)
            _claude_fill_form(page, profile)
            _uncheck_optionals(page)  # second pass — some appear after fills

        time.sleep(0.3)
        result = _advance(page)
        print(f"      [apply] Step {step + 1}: {result}")

        # Detect stuck loop — same non-submit button clicked 3 times in a row
        if result == last_result and result not in ("submit", "stuck"):
            repeat_count += 1
            if repeat_count >= 3:
                err_el = page.query_selector(".artdeco-inline-feedback--error")
                err_msg = err_el.inner_text()[:120] if err_el else "unknown"
                print(f"      [apply] Stuck on '{result}' × {repeat_count} — form error: {err_msg}")
                return False
        else:
            repeat_count = 0
        last_result = result

        if result == "submit":
            time.sleep(2)
            # Check for form validation errors
            err = page.query_selector(
                ".artdeco-inline-feedback--error, .jobs-easy-apply-form-element__error"
            )
            if err:
                print(f"      [apply] Blocked by error: {err.inner_text()[:120]}")
                return False
            return True

        if result == "stuck":
            print(f"      [apply] No navigation button found at step {step + 1}")
            return False

    print("      [apply] Reached step limit without submitting")
    return False


# ===========================================================================
# Public API
# ===========================================================================

def attempt_easy_apply(page: Page, job: dict, profile: dict, cv_path: Path | None) -> bool:
    """
    Attempt to submit a LinkedIn Easy Apply for one job.
    Returns True if the application was submitted successfully.
    """
    title = job.get("title", "?")
    company = job.get("company", "?")
    print(f"    [apply] {title} @ {company}")

    try:
        # Navigate and wait only until DOM is ready, then wait for a job element
        page.goto(job["url"], wait_until="domcontentloaded", timeout=30000)

        for sel in [
            ".jobs-apply-button",
            ".jobs-unified-top-card__job-title",
            ".job-details-jobs-unified-top-card__job-title",
            ".jobs-details__main-content",
        ]:
            try:
                page.wait_for_selector(sel, timeout=8000)
                break
            except Exception:
                pass

        time.sleep(0.5)
        print(f"      [apply] URL: {page.url[:90]}")

        # Find and click Easy Apply via JS.
        # LinkedIn renders this as either a <button> or an <a> tag depending
        # on the page version, so we search both by aria-label and by text.
        result = page.evaluate("""() => {
            // 1. aria-label match on any element (button OR anchor)
            let el = document.querySelector('[aria-label*="Easy Apply"]');
            // 2. text match across buttons and anchors
            if (!el) {
                el = [...document.querySelectorAll('button, a')].find(e =>
                    e.innerText && e.innerText.trim().includes('Easy Apply')
                );
            }
            if (!el || el.disabled) return null;
            // Detect already-applied state
            if (el.innerText && el.innerText.toLowerCase().includes('applied') &&
                !el.innerText.toLowerCase().includes('easy apply')) {
                return 'already_applied';
            }
            el.click();
            return el.innerText ? el.innerText.trim() : 'clicked';
        }""")

        if result is None:
            print("    [apply] No Easy Apply button — skipping")
            return False
        if result == "already_applied":
            print("    [apply] Already applied (button says Applied) — skipping")
            return False

        time.sleep(2)

        submitted = _run_modal(page, profile, cv_path)

        if not submitted:
            _close_modal(page)
            time.sleep(1)

        return submitted

    except Exception as exc:
        print(f"    [apply] Error: {exc}")
        try:
            _close_modal(page)
        except Exception:
            pass
        return False


